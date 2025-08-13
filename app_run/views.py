from typing import ClassVar

from app_athletes.models import AthleteInfo
from app_challenges.models import Challenge
from app_positions.models import Position
from app_subscribe.models import Subscribe
from collectible_items.serializers import CollectibleItemSerializer
from django.db.models import Avg, Max, Min, Sum
from django.shortcuts import get_object_or_404
from django_filters.rest_framework import DjangoFilterBackend
from geopy import distance
from openpyxl import load_workbook
from project_run.settings.base import (
    CHALLENGE_2_KILOMETERS_IN_10_MINUTES,
    CHALLENGE_50_KILOMETERS_RUNS,
    CHALLENGE_DO_10_RUNS,
    COMPANY_NAME,
    CONTACTS,
    SLOGAN,
)
from rest_framework import status
from rest_framework.decorators import api_view
from rest_framework.filters import OrderingFilter
from rest_framework.pagination import PageNumberPagination
from rest_framework.request import Request
from rest_framework.response import Response
from rest_framework.views import APIView
from rest_framework.viewsets import ModelViewSet

from app_run.models import Run
from app_run.serializers import (
    RunSerializer,
)


@api_view(["GET"])
def company_details(request: Request) -> Response:  # noqa: ARG001
    return Response(
        {
            "company_name": COMPANY_NAME,
            "slogan": SLOGAN,
            "contacts": CONTACTS,
        },
    )


@api_view(["POST"])
def upload_file(request: Request) -> Response:
    file = request.FILES.get("file")
    wb = load_workbook(file)
    sheet = wb.active
    headers = [
        "name",
        "uid",
        "value",
        "latitude",
        "longitude",
        "url",
    ]
    rows = list(sheet.iter_rows(values_only=True, min_row=2))
    error_data = []
    for _idx, row in enumerate(rows, start=2):
        data = dict(zip(headers, row, strict=False))
        serializer = CollectibleItemSerializer(data=data)
        if serializer.is_valid():
            serializer.save()
        else:
            error_data.append(list(row))
    return Response(data=error_data)


class RunPagination(PageNumberPagination):
    page_size_query_param = "size"
    max_page_size = 50


class RunViewSet(ModelViewSet):
    queryset = Run.objects.all().prefetch_related("athlete")
    serializer_class = RunSerializer
    filter_backends: ClassVar[list] = [
        DjangoFilterBackend,
        OrderingFilter,
    ]
    filterset_fields: ClassVar[list[str]] = [
        "status",
        "athlete",
    ]
    ordering_fields: ClassVar[list[str]] = [
        "created_at",
    ]
    pagination_class = RunPagination


class RunAPIStartView(APIView):
    def post(self, request: Request, run_id: int) -> Response:  # noqa: ARG002
        run = get_object_or_404(Run, id=run_id)
        if run.status == Run.INITIAL:
            run.status = Run.IN_PROGRESS
            run.save()
            return Response(
                {"message": "Run started"},
                status=status.HTTP_200_OK,
            )
        return Response(
            {"message": "Run already started or finished"},
            status=status.HTTP_400_BAD_REQUEST,
        )


class RunAPIStopView(APIView):
    def post(self, request: Request, run_id: int) -> Response:  # noqa: ARG002
        run_distance = 0
        run = get_object_or_404(
            Run,
            id=run_id,
        )
        if run.status == Run.IN_PROGRESS:
            run.status = Run.FINISHED
            positions = Position.objects.filter(run=run)
            for i in range(len(positions) - 1):
                run_distance += distance.distance(
                    (positions[i].latitude, positions[i].longitude),
                    (positions[i + 1].latitude, positions[i + 1].longitude),
                ).km
            run.distance = round(run_distance, 3)
            run.save()
            finished_run_count = Run.objects.filter(
                status=Run.FINISHED,
                athlete=run.athlete.id,
            ).count()
            if finished_run_count == 10:
                challenge = Challenge(
                    full_name=CHALLENGE_DO_10_RUNS,
                    athlete=AthleteInfo.objects.get(user_id=run.athlete),
                )
                challenge.save()
            total_distance = Run.objects.filter(athlete=run.athlete).aggregate(
                Sum("distance"),
            )
            if total_distance.get("distance__sum") >= 50:
                challenge = Challenge(
                    full_name=CHALLENGE_50_KILOMETERS_RUNS,
                    athlete=AthleteInfo.objects.get(user_id=run.athlete),
                )
                challenge.save()
            if (
                positions.aggregate(Sum("speed")).get("speed__sum")
                and positions.count() != 0
            ):
                run.speed = round(
                    positions.aggregate(Sum("speed")).get("speed__sum")
                    / positions.count(),
                    2,
                )
                run.save()
            first_position_time = positions.aggregate(Min("date_time")).get(
                "date_time__min",
            )
            last_position_time = positions.aggregate(Max("date_time")).get(
                "date_time__max",
            )
            if first_position_time and last_position_time:
                diff_seconds = (last_position_time - first_position_time).seconds
                run.run_time_seconds = int(diff_seconds)
                run.save()
            if run.distance >= 2 and run.run_time_seconds / 60 <= 10:
                challenge = Challenge(
                    full_name=CHALLENGE_2_KILOMETERS_IN_10_MINUTES,
                    athlete=AthleteInfo.objects.get(user_id=run.athlete),
                )
                challenge.save()
            return Response(
                {"message": "Run stopped"},
                status=status.HTTP_200_OK,
            )

        return Response(
            {"message": "Run not started or already finished"},
            status=status.HTTP_400_BAD_REQUEST,
        )


class CoachAnalyticsView(APIView):
    def get(self, request: Request, coach_id: int) -> Response:  # noqa: ARG002
        subscribe = Subscribe.objects.filter(coach=coach_id).select_related("athlete")
        longest_run = (
            subscribe.annotate(longest=Max("athlete__run__distance"))
            .order_by("-longest")
            .first()
        )
        total_run = (
            subscribe.annotate(sum=Sum("athlete__run__distance"))
            .order_by("-sum")
            .first()
        )
        avg_speed = (
            subscribe.annotate(avg_speed=Avg("athlete__run__speed"))
            .order_by("-avg_speed")
            .first()
        )
        return Response(
            {
                "longest_run_user": longest_run.athlete_id,
                "longest_run_value": longest_run.longest,
                "total_run_user": total_run.athlete_id,
                "total_run_value": total_run.sum,
                "speed_avg_user": avg_speed.athlete_id,
                "speed_avg_value": avg_speed.avg_speed,
            },
        )
